"""
Utilities for report generation and formatting.
"""
from django.db.models import Sum, Count, Avg
from datetime import datetime, timedelta
from decimal import Decimal


class ReportUtils:
    """Utility class for report operations."""
    
    @staticmethod
    def format_currency(amount):
        """Format amount as currency."""
        if amount is None:
            return "0.00"
        return f"{amount:,.2f}"
    
    @staticmethod
    def calculate_percentage(part, total):
        """Calculate percentage."""
        if total == 0:
            return 0
        return round((part / total) * 100, 2)
    
    @staticmethod
    def get_date_ranges(period):
        """Get start and end dates for a period."""
        today = datetime.now().date()
        
        if period == 'today':
            start_date = today
            end_date = today + timedelta(days=1)
        elif period == 'yesterday':
            start_date = today - timedelta(days=1)
            end_date = today
        elif period == 'week':
            start_date = today - timedelta(days=7)
            end_date = today + timedelta(days=1)
        elif period == 'month':
            start_date = today - timedelta(days=30)
            end_date = today + timedelta(days=1)
        elif period == 'year':
            start_date = today - timedelta(days=365)
            end_date = today + timedelta(days=1)
        else:
            start_date = None
            end_date = None
        
        return start_date, end_date
    
    @staticmethod
    def generate_monthly_data(queryset, date_field, value_field, months=12):
        """Generate monthly aggregated data."""
        from django.utils import timezone
        
        today = timezone.now().date()
        monthly_data = []
        
        for i in range(months):
            month_start = today.replace(day=1) - timedelta(days=30*i)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            # Filter queryset for this month
            month_filter = {
                f'{date_field}__date__gte': month_start,
                f'{date_field}__date__lte': month_end
            }
            
            month_queryset = queryset.filter(**month_filter)
            
            total = month_queryset.aggregate(
                total=Sum(value_field)
            )['total'] or Decimal('0.00')
            
            count = month_queryset.count()
            
            monthly_data.append({
                'month': month_start.strftime('%Y-%m'),
                'month_name': month_start.strftime('%b %Y'),
                'total': total,
                'count': count,
                'average': total / count if count > 0 else Decimal('0.00')
            })
        
        return monthly_data
    
    @staticmethod
    def generate_weekly_data(queryset, date_field, value_field, weeks=8):
        """Generate weekly aggregated data."""
        from django.utils import timezone
        
        today = timezone.now().date()
        weekly_data = []
        
        for i in range(weeks):
            week_start = today - timedelta(days=today.weekday() + (7 * i))
            week_end = week_start + timedelta(days=6)
            
            # Filter queryset for this week
            week_filter = {
                f'{date_field}__date__gte': week_start,
                f'{date_field}__date__lte': week_end
            }
            
            week_queryset = queryset.filter(**week_filter)
            
            total = week_queryset.aggregate(
                total=Sum(value_field)
            )['total'] or Decimal('0.00')
            
            count = week_queryset.count()
            
            weekly_data.append({
                'week_start': week_start.strftime('%Y-%m-%d'),
                'week_end': week_end.strftime('%Y-%m-%d'),
                'week_label': f"{week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')}",
                'total': total,
                'count': count
            })
        
        return weekly_data
    
    @staticmethod
    def generate_daily_data(queryset, date_field, value_field, days=30):
        """Generate daily aggregated data."""
        from django.utils import timezone
        
        today = timezone.now().date()
        daily_data = []
        
        for i in range(days):
            date = today - timedelta(days=i)
            
            # Filter queryset for this day
            date_filter = {
                f'{date_field}__date': date
            }
            
            day_queryset = queryset.filter(**date_filter)
            
            total = day_queryset.aggregate(
                total=Sum(value_field)
            )['total'] or Decimal('0.00')
            
            count = day_queryset.count()
            
            daily_data.append({
                'date': date.strftime('%Y-%m-%d'),
                'day_name': date.strftime('%a'),
                'total': total,
                'count': count
            })
        
        # Reverse to show oldest first
        daily_data.reverse()
        return daily_data
    
    @staticmethod
    def export_to_excel(data, filename):
        """Export data to Excel file."""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Write headers
        if data and isinstance(data, list) and data[0]:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, row in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header)
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save file
        wb.save(filename)
        return filename